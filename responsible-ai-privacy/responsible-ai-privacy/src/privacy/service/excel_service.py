'''
MIT license https://opensource.org/licenses/MIT Copyright 2024 Infosys Ltd

Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated documentation files (the "Software"), to deal in the Software without restriction, including without limitation the rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and to permit persons to whom the Software is furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
'''

import base64
import json
import io, base64
from PIL import Image
import requests
import pandas as pd
from privacy.mappers.mappers import *
from privacy.service.textPrivacy import *
import os
import openpyxl
from openpyxl.reader.excel import load_workbook
import xlsxwriter
import re 

from privacy.config.logger import CustomLogger
from privacy.service.textPrivacy import TextPrivacy
log = CustomLogger()

class AttributeDict(dict):
    __getattr__ = dict.__getitem__
    __setattr__ = dict.__setitem__
    __delattr__ = dict.__delitem__

import shutil



def path_check(safe_path):
    # Ensure the path is valid and does not contain any illegal characters
    if re.match(r'^[\w\-\\\/\s.]+$', str(safe_path)):
        return safe_path
    else:
        raise ValueError(f"Invalid path: {safe_path}")

def is_safe_path(basedir, path, follow_symlinks=True):
    # Normalize the path
    normalized_path = os.path.normpath(path)
    
    # Resolve symbolic links
    if follow_symlinks:
        real_path = os.path.realpath(normalized_path)
    else:
        real_path = os.path.abspath(normalized_path)
    
    # Check if the resolved path starts with the base directory
    return os.path.commonpath([real_path, basedir]) == basedir

def sanitize_filename(filename):
    # Allow only specific characters: alphanumeric, underscores, hyphens, and dots
    allowed_chars = re.compile(r'^[a-zA-Z0-9_.-]+$')
    if allowed_chars.match(filename):
        return filename
    else:
        raise ValueError(f"Invalid filename: {filename}")

class Excel:

    @staticmethod
    def excelanonymize(payload):
        payload = AttributeDict(payload)

        # Create a secure temporary directory
        temp_dir = "temp_excel_files"
        os.makedirs(temp_dir, exist_ok=True)

        # Generate a unique filename
        temp_filename = f"temp_{os.urandom(16).hex()}.xlsx"
        temp_filename = sanitize_filename(temp_filename)  # Sanitize the filename
        temp_filepath = os.path.join(temp_dir, temp_filename)

        # Validate the temp_filepath using path_check
        path_check(temp_filepath)

        # Ensure the temp_filepath is safe
        if not is_safe_path(temp_dir, temp_filepath):
            raise ValueError(f"Unsafe file path detected: {temp_filepath}")

        try:
            with open(temp_filepath, "wb") as f:
                shutil.copyfileobj(payload.excel.file, f)

            wrkbk = openpyxl.load_workbook(temp_filepath)
            sh = wrkbk.active

            x = ""
            s = ""
            row = 0
            col = 0
            output_filename = "ExcelOutput.xlsx"
            output_filepath = os.path.join(temp_dir, output_filename)
            workbook = xlsxwriter.Workbook(output_filepath)
            worksheet = workbook.add_worksheet()
            r = 0
            list = []
            i = 0

            for row in sh.iter_rows(min_row=1, min_col=1):
                for cell in row:
                    cell_coor = cell.coordinate
                    print("cell", cell.coordinate)
                    print("cell_value=====", cell.value)
                    payload1 = {"inputText": str(cell.value), "exclusionList": None, "portfolio": None, "account": None, "fakeData": False}
                    payload1 = AttributeDict(payload1)

                    temp = TextPrivacy.anonymize(payload1)
                    print("response===", temp.anonymizedText)

                    print("Old cell value===", str(cell.value))
                    sh[str(cell_coor)] = temp.anonymizedText
                    print("New cell value===", str(cell.value))

            wrkbk.save(filename=output_filepath)
            print("temp====", s)
            print("x=====", x)
            print("Workbook==", workbook)
            workbook.close()
            return output_filepath

        except Exception as e:
            log.error(f"Error in excelanonymize: {e}")
            raise

        finally:
            if os.path.exists(temp_filepath):
                try:
                    os.remove(temp_filepath)
                except Exception as e:
                    log.error(f"Error removing temporary file: {str(e)}")

    @staticmethod
    def createExcel(s, x, row, col, worksheet):
        print("row187===", row)

        for x in s.split(';*'):
            worksheet.write(row, col, x)
            col += 1
